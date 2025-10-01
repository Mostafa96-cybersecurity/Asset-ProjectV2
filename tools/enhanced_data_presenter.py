#!/usr/bin/env python3
"""
🎨 Enhanced Data Presentation & Smart Duplicate Management
=========================================================

This tool creates a professional, organized data presentation system
with beautiful formatting and intelligent duplicate handling.
"""

import pandas as pd
from datetime import datetime
import re

class EnhancedDataPresenter:
    """Professional data presentation with smart formatting"""
    
    def __init__(self):
        self.duplicate_counter = {}
        
    def format_storage_details(self, storage_data):
        """Convert storage data to organized format"""
        if not storage_data or storage_data == "N/A":
            return "Storage: Not Available"
        
        # Handle different storage formats
        if "Disk" in storage_data:
            # Already formatted correctly
            return storage_data.replace('\n', ' | ')
        
        # Format raw storage data
        formatted_storage = []
        disk_count = 1
        
        # Extract storage values
        storage_values = re.findall(r'(\d+\.?\d*)\s*GB', storage_data)
        for value in storage_values:
            formatted_storage.append(f"Disk {disk_count} = {value}GB")
            disk_count += 1
        
        return " | ".join(formatted_storage) if formatted_storage else storage_data

    def format_cpu_details(self, cpu_data):
        """Format CPU information professionally"""
        if not cpu_data or cpu_data == "N/A":
            return "CPU: Not Available"
        
        # Clean up CPU formatting
        cpu_lines = cpu_data.split('\n')
        unique_cpus = list(set(line.strip() for line in cpu_lines if line.strip()))
        
        if len(unique_cpus) == 1:
            return unique_cpus[0]
        else:
            return f"{len(unique_cpus)}x {unique_cpus[0] if unique_cpus else 'Unknown CPU'}"

    def handle_smart_duplicates(self, devices):
        """Smart duplicate handling without data loss"""
        print("🔧 SMART DUPLICATE MANAGEMENT")
        print("=" * 50)
        
        fingerprint_map = {}
        processed_devices = []
        duplicate_info = {}
        
        for device in devices:
            # Create fingerprint for duplicate detection
            fingerprint = self.create_device_fingerprint(device)
            
            if fingerprint in fingerprint_map:
                # Found duplicate - merge intelligently
                original_device = fingerprint_map[fingerprint]
                merged_device = self.merge_device_data(original_device, device)
                
                # Track duplicate info
                hostname = device.get('Hostname', 'Unknown')
                if fingerprint not in duplicate_info:
                    duplicate_info[fingerprint] = {
                        'original': original_device.get('Hostname', 'Unknown'),
                        'duplicates': []
                    }
                duplicate_info[fingerprint]['duplicates'].append(hostname)
                
                # Update the device in processed list
                for i, processed in enumerate(processed_devices):
                    if self.create_device_fingerprint(processed) == fingerprint:
                        processed_devices[i] = merged_device
                        break
                
                print(f"   🔄 DUPLICATE MERGED: {hostname}")
                print(f"      Original: {original_device.get('Hostname')}")
                print(f"      Merged into: {merged_device.get('Hostname')}")
                
            else:
                # New unique device
                fingerprint_map[fingerprint] = device
                processed_devices.append(device)
                print(f"   ✅ NEW DEVICE: {device.get('Hostname', 'Unknown')}")
        
        # Show duplicate summary
        if duplicate_info:
            print(f"\n📊 DUPLICATE SUMMARY:")
            for fp, info in duplicate_info.items():
                print(f"   Original: {info['original']}")
                print(f"   Merged: {', '.join(info['duplicates'])}")
                print(f"   Result: Enhanced data with no loss")
        
        return processed_devices

    def create_device_fingerprint(self, device):
        """Create fingerprint for duplicate detection"""
        # Use Serial Number as primary identifier
        serial = device.get('Serial Number', device.get('SN', '')).strip()
        if serial and len(serial) > 4:
            return f"SN:{serial}"
        
        # Fallback to MAC Address
        mac = device.get('MAC Address', '').strip()
        if mac:
            return f"MAC:{mac}"
        
        # Fallback to Hostname + IP
        hostname = device.get('Hostname', '').strip().lower()
        ip = device.get('IP Address', '').strip()
        if hostname and ip:
            return f"HOST:{hostname}@{ip}"
        
        return f"UNIQUE:{datetime.now().timestamp()}"

    def merge_device_data(self, original, duplicate):
        """Intelligently merge duplicate device data"""
        merged = original.copy()
        
        for key, value in duplicate.items():
            if key not in merged or not merged[key] or merged[key] == "N/A":
                # Add missing data
                merged[key] = value
            elif value and value != "N/A" and value != merged[key]:
                # Handle conflicts intelligently
                if key == 'Hostname':
                    # Keep the more descriptive hostname
                    merged[key] = value if len(value) > len(merged[key]) else merged[key]
                elif key == 'Working User':
                    # Combine users if different
                    if value not in merged[key]:
                        merged[key] = f"{merged[key]} | {value}"
                elif key in ['Storage (Hard Disk)', 'CPU Processor']:
                    # Combine technical specs
                    merged[key] = f"{merged[key]} | {value}"
                elif key == 'Monitor':
                    # Combine monitor info
                    if value not in merged[key]:
                        merged[key] = f"{merged[key]} | {value}"
        
        # Add merge timestamp
        merged['_Last_Merged'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        merged['_Merge_Source'] = 'Smart Duplicate Handler'
        
        return merged

    def create_professional_presentation(self, devices):
        """Create professionally formatted data presentation"""
        print("\n🎨 PROFESSIONAL DATA PRESENTATION")
        print("=" * 80)
        
        # Process each device
        formatted_devices = []
        
        for i, device in enumerate(devices, 1):
            print(f"\n📋 DEVICE {i}: {device.get('Hostname', 'Unknown')}")
            print("-" * 60)
            
            # Format and display key information
            formatted_device = {
                'ID': f"DEV-{i:03d}",
                'Hostname': device.get('Hostname', 'N/A'),
                'IP Address': device.get('IP Address', 'N/A'),
                'Working User': device.get('Working User', 'N/A'),
                'Domain': device.get('Domain', 'N/A'),
                'Device Model': device.get('Device Model', 'N/A'),
                'Manufacturer': device.get('Manufacturer', 'N/A'),
                'Serial Number': device.get('Serial Number', 'N/A'),
                'MAC Address': device.get('MAC Address', 'N/A'),
                'OS Name and Version': device.get('OS Name and Version', 'N/A'),
                'Installed RAM (GB)': f"{device.get('Installed RAM (GB)', 'N/A')} GB",
                'CPU Processor': self.format_cpu_details(device.get('CPU Processor', 'N/A')),
                'Storage Details': self.format_storage_details(device.get('Storage (Hard Disk)', 'N/A')),
                'Monitor': device.get('Monitor', 'N/A'),
                'Graphics Card': device.get('Graphics Card', 'N/A'),
                'Status': '🟢 Active' if device.get('IP Address', 'N/A') != 'N/A' else '🔴 Offline',
                'Last Updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # Display formatted information
            for key, value in formatted_device.items():
                if key in ['CPU Processor', 'Storage Details']:
                    print(f"   {key:<20}: {value}")
                else:
                    print(f"   {key:<20}: {value}")
            
            formatted_devices.append(formatted_device)
        
        return formatted_devices

    def generate_excel_output(self, devices, filename="Enhanced_Asset_Report.xlsx"):
        """Generate professional Excel output with formatting"""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # Create DataFrame
            df = pd.DataFrame(devices)
            
            # Create Excel writer
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Asset Report', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Asset Report']
                
                # Define styles
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                data_fill_even = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
                data_fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                # Style headers
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Style data rows
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                    fill = data_fill_even if row_num % 2 == 0 else data_fill_odd
                    for cell in row:
                        cell.fill = fill
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            print(f"\n📊 EXCEL REPORT GENERATED: {filename}")
            return filename
            
        except ImportError:
            print("⚠️ openpyxl not installed. Using basic CSV export...")
            csv_filename = filename.replace('.xlsx', '.csv')
            pd.DataFrame(devices).to_csv(csv_filename, index=False)
            return csv_filename


def demonstrate_enhanced_presentation():
    """Demonstrate enhanced data presentation with your sample data"""
    
    # Your sample data
    sample_data = [
        {
            'IP Address': '10.0.23.99',
            'Hostname': 'MHQ-ENG-SAHMED',
            'Working User': r'SQUARE\sara.ahmed',
            'Domain': 'square.local',
            'Device Model': 'Precision Tower 7810',
            'MAC Address': '50:9A:4C:42:D4:09',
            'OS Name and Version': 'Microsoft Windows 10 Pro',
            'Installed RAM (GB)': '31.92',
            'CPU Processor': 'Intel(R) Xeon(R) CPU E5-2680 v4 @ 2.40GHz\nIntel(R) Xeon(R) CPU E5-2680 v4 @ 2.40GHz',
            'Storage (Hard Disk)': 'Disk 1 = 232.88GB\nDisk 2 = 931.51GB',
            'Manufacturer': 'Dell Inc.',
            'Serial Number': 'BD35LH2',
            'Monitor': 'HP EliteDisplay E221c Webcam LED Backlit Monitor',
            'Graphics Card': 'NVIDIA GeForce GTX 1660 Ti',
            'Installed Key': 'N/A'
        },
        {
            'IP Address': '10.0.21.246',
            'Hostname': 'MHQ-ENG-HMETWAL',
            'Working User': r'SQUARE\hager.metwally',
            'Domain': 'square.local',
            'Device Model': 'Precision Tower 7810',
            'MAC Address': '50:9A:4C:45:25:D5',
            'OS Name and Version': 'Microsoft Windows 10 Pro',
            'Installed RAM (GB)': '31.92',
            'CPU Processor': 'Intel(R) Xeon(R) CPU E5-2680 v4 @ 2.40GHz\nIntel(R) Xeon(R) CPU E5-2680 v4 @ 2.40GHz',
            'Storage (Hard Disk)': 'Disk 1 = 465.76GB',
            'Manufacturer': 'Dell Inc.',
            'Serial Number': '7QHHKH2',
            'Monitor': 'HP EliteDisplay E231 LED Backlit Monitor',
            'Graphics Card': 'NVIDIA GeForce GTX 1660 Ti',
            'Installed Key': 'N/A'
        },
        {
            'IP Address': '10.0.23.76',
            'Hostname': 'MHQ-BIM-MGAD',
            'Working User': r'SQUARE\mohamed.gad',
            'Domain': 'square.local',
            'Device Model': 'Precision Tower 7910',
            'MAC Address': 'D8:9E:F3:45:A6:20',
            'OS Name and Version': 'Microsoft Windows 10 Pro',
            'Installed RAM (GB)': '31.92',
            'CPU Processor': 'Intel(R) Xeon(R) CPU E5-2697 v4 @ 2.30GHz\nIntel(R) Xeon(R) CPU E5-2697 v4 @ 2.30GHz',
            'Storage (Hard Disk)': 'Disk 1 = 232.88GB',
            'Manufacturer': 'Dell Inc.',
            'Serial Number': '17XCCP2',
            'Monitor': 'Generic PnP Monitor',
            'Graphics Card': 'NVIDIA GeForce RTX 3060',
            'Installed Key': 'N/A'
        },
        # Add a duplicate for testing
        {
            'IP Address': '10.0.23.99',  # Same IP as first device
            'Hostname': 'MHQ-ENG-SAHMED-DUPLICATE',  # Slightly different hostname
            'Working User': r'SQUARE\sara.ahmed',
            'Domain': 'square.local',
            'Device Model': 'Precision Tower 7810',
            'MAC Address': '50:9A:4C:42:D4:09',  # Same MAC as first device
            'OS Name and Version': 'Microsoft Windows 10 Pro',
            'Installed RAM (GB)': '31.92',
            'CPU Processor': 'Intel(R) Xeon(R) CPU E5-2680 v4 @ 2.40GHz',
            'Storage (Hard Disk)': 'Disk 3 = 1TB',  # Additional storage info
            'Manufacturer': 'Dell Inc.',
            'Serial Number': 'BD35LH2',  # Same serial - this makes it a duplicate
            'Monitor': 'Additional Monitor Info',
            'Graphics Card': 'NVIDIA GeForce GTX 1660 Ti',
            'Installed Key': 'Windows 10 Pro'
        }
    ]
    
    # Create presenter
    presenter = EnhancedDataPresenter()
    
    # Handle duplicates smartly
    unique_devices = presenter.handle_smart_duplicates(sample_data)
    
    # Create professional presentation
    formatted_devices = presenter.create_professional_presentation(unique_devices)
    
    # Generate Excel report
    excel_file = presenter.generate_excel_output(formatted_devices)
    
    print(f"\n🎉 ENHANCED PRESENTATION COMPLETE!")
    print(f"📁 Total Devices: {len(sample_data)} → {len(unique_devices)} (after smart duplicate handling)")
    print(f"📊 Excel Report: {excel_file}")
    print(f"✅ Professional formatting applied")
    print(f"🛡️ Smart duplicate management completed")


if __name__ == "__main__":
    demonstrate_enhanced_presentation()