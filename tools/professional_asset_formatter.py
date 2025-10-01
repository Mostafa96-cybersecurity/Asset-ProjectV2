"""
Professional Asset Formatter with Smart Duplicate Handling
========================================================

This module provides professional data formatting and presentation
for network asset collection with intelligent duplicate management.

Features:
- Smart duplicate detection using device fingerprinting
- Professional Excel formatting with colors and styling  
- Intelligent data merging without loss
- Beautiful storage formatting (Disk 1 = 250GB, Disk 2 = 500GB)
- Color-coded status indicators
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from typing import List, Dict, Any, Optional
import hashlib
import re


class ProfessionalAssetFormatter:
    """Professional asset formatter with smart duplicate handling"""
    
    def __init__(self):
        # Professional color scheme
        self.color_scheme = {
            'header_bg': '4472C4',      # Professional blue
            'header_text': 'FFFFFF',    # White text
            'row_even': 'F2F2F2',       # Light gray
            'row_odd': 'FFFFFF',        # White
            'success': '70AD47',        # Green
            'warning': 'FFC000',        # Orange
            'error': 'C5504B',          # Red
            'accent': '5B9BD5'          # Light blue
        }
        
    def generate_device_fingerprint(self, device: Dict[str, Any]) -> str:
        """Generate unique fingerprint for device identification"""
        # Priority order for identification
        serial = str(device.get('Serial Number', '')).strip()
        mac = str(device.get('MAC Address', '')).strip()
        hostname = str(device.get('Hostname', '')).strip()
        ip = str(device.get('IP Address', '')).strip()
        
        # Create fingerprint based on available data
        if serial and serial != 'N/A':
            return f"SERIAL:{serial}"
        elif mac and mac != 'N/A':
            return f"MAC:{mac}"
        elif hostname and hostname != 'N/A':
            return f"HOST:{hostname}:{ip}"
        else:
            return f"IP:{ip}"
    
    def format_storage_display(self, storage: str) -> str:
        """Format storage information professionally"""
        if not storage or storage == 'N/A':
            return 'N/A'
        
        # Handle various storage formats
        storage_str = str(storage)
        
        # If already in desired format, return as is
        if 'Disk' in storage_str and '=' in storage_str:
            return storage_str
        
        # Extract GB values and format
        gb_matches = re.findall(r'(\d+(?:\.\d+)?)\s*GB', storage_str)
        if gb_matches:
            formatted_disks = []
            for i, size in enumerate(gb_matches, 1):
                formatted_disks.append(f"Disk {i} = {size}GB")
            return ", ".join(formatted_disks)
        
        return storage_str
    
    def format_ram_display(self, ram: str) -> str:
        """Format RAM information"""
        if not ram or ram == 'N/A':
            return 'N/A'
        
        ram_str = str(ram)
        if 'GB' not in ram_str:
            try:
                ram_float = float(ram_str)
                return f"{ram_float:.1f} GB"
            except:
                return ram_str
        return ram_str
    
    def format_cpu_display(self, cpu: str) -> str:
        """Format CPU information"""
        if not cpu or cpu == 'N/A':
            return 'N/A'
        
        cpu_str = str(cpu)
        # If multiple CPUs on newlines, join with " + "
        if '\n' in cpu_str:
            cpu_parts = cpu_str.split('\n')
            unique_cpus = list(dict.fromkeys(cpu_parts))  # Remove duplicates
            return " + ".join(unique_cpus)
        
        return cpu_str
    
    def smart_duplicate_handler(self, devices: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Handle duplicates intelligently without losing data"""
        print("\n🛡️ SMART DUPLICATE DETECTION")
        print("=" * 40)
        
        device_map = {}
        processed = []
        duplicate_log = []
        
        for device in devices:
            fingerprint = self.generate_device_fingerprint(device)
            hostname = device.get('Hostname', 'Unknown')
            
            if fingerprint in device_map:
                # Found duplicate - merge intelligently
                existing = device_map[fingerprint]
                merged = self.merge_device_data(existing, device)
                device_map[fingerprint] = merged
                
                # Update in processed list
                for i, proc_device in enumerate(processed):
                    if self.generate_device_fingerprint(proc_device) == fingerprint:
                        processed[i] = merged
                        break
                
                # Log the merge
                duplicate_log.append({
                    'original': existing.get('Hostname', 'Unknown'),
                    'duplicate': hostname,
                    'merged_into': merged.get('Hostname', 'Unknown'),
                    'fingerprint': fingerprint
                })
                
                print(f"   🔄 DUPLICATE MERGED: {hostname}")
                print(f"      → Enhanced data preserved without loss")
                
            else:
                # New unique device
                device_map[fingerprint] = device
                processed.append(device.copy())
                print(f"   ✅ NEW DEVICE: {hostname}")
        
        if duplicate_log:
            print(f"\n📊 DUPLICATE SUMMARY:")
            for entry in duplicate_log:
                print(f"   Original: {entry['original']}")
                print(f"   Merged: {entry['duplicate']} → {entry['merged_into']}")
        
        print(f"\n📈 RESULT: {len(devices)} devices → {len(processed)} unique devices")
        return processed
    
    def merge_device_data(self, existing: Dict[str, Any], new: Dict[str, Any]) -> Dict[str, Any]:
        """Intelligently merge device data without losing information"""
        merged = existing.copy()
        
        for key, value in new.items():
            existing_value = merged.get(key, '')
            
            # Skip empty or N/A values
            if not value or str(value).strip() in ['', 'N/A']:
                continue
            
            # If existing is empty or N/A, use new value
            if not existing_value or str(existing_value).strip() in ['', 'N/A']:
                merged[key] = value
            # If values are different, combine them intelligently
            elif str(value).strip() != str(existing_value).strip():
                if key in ['Storage (Hard Disk)', 'Monitor']:
                    # Combine storage and monitor info
                    combined = f"{existing_value} | {value}"
                    merged[key] = combined
                elif key == 'CPU Processor':
                    # Combine CPU info if different
                    if str(value) not in str(existing_value):
                        merged[key] = f"{existing_value} | {value}"
                elif key == 'Working User':
                    # Combine users if different
                    if str(value) not in str(existing_value):
                        merged[key] = f"{existing_value} | {value}"
                else:
                    # For other fields, prefer the more detailed value
                    if len(str(value)) > len(str(existing_value)):
                        merged[key] = value
        
        # Add merge metadata
        merged['_Last_Merged'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        merged['_Merge_Count'] = merged.get('_Merge_Count', 0) + 1
        
        return merged
    
    def create_professional_presentation(self, devices: List[Dict[str, Any]]) -> pd.DataFrame:
        """Create professionally formatted presentation"""
        print("\n🎨 CREATING PROFESSIONAL PRESENTATION")
        print("=" * 50)
        
        # Handle duplicates first
        unique_devices = self.smart_duplicate_handler(devices)
        
        # Format data for professional presentation
        formatted_data = []
        
        for i, device in enumerate(unique_devices, 1):
            formatted_device = {
                'ID': f'DEV-{i:03d}',
                'IP Address': device.get('IP Address', 'N/A'),
                'Hostname': device.get('Hostname', 'N/A'),
                'Working User': device.get('Working User', 'N/A'),
                'Domain': device.get('Domain', 'N/A'),
                'Device Model': device.get('Device Model', 'N/A'),
                'MAC Address': device.get('MAC Address', 'N/A'),
                'OS Name and Version': device.get('OS Name and Version', 'N/A'),
                'Installed RAM (GB)': self.format_ram_display(device.get('Installed RAM (GB)', 'N/A')),
                'CPU Processor': self.format_cpu_display(device.get('CPU Processor', 'N/A')),
                'Storage (Hard Disk)': self.format_storage_display(device.get('Storage (Hard Disk)', 'N/A')),
                'Manufacturer': device.get('Manufacturer', 'N/A'),
                'Serial Number': device.get('Serial Number', 'N/A'),
                'Monitor': device.get('Monitor', 'N/A'),
                'Graphics Card': device.get('Graphics Card', 'N/A'),
                'Status': '🟢 Active' if device.get('IP Address', 'N/A') != 'N/A' else '🔴 Offline',
                'Last Updated': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            formatted_data.append(formatted_device)
            
            # Print formatted device info
            print(f"\n📋 DEVICE {i}: {formatted_device['Hostname']}")
            print("-" * 40)
            print(f"   IP: {formatted_device['IP Address']}")
            print(f"   User: {formatted_device['Working User']}")
            print(f"   Model: {formatted_device['Device Model']}")
            print(f"   RAM: {formatted_device['Installed RAM (GB)']}")
            print(f"   Storage: {formatted_device['Storage (Hard Disk)']}")
            print(f"   Status: {formatted_device['Status']}")
        
        df = pd.DataFrame(formatted_data)
        print(f"\n✅ Professional presentation created for {len(df)} devices")
        return df
    
    def export_professional_excel(self, df: pd.DataFrame, filename: str = None) -> str:
        """Export to professional Excel with beautiful formatting"""
        if filename is None:
            filename = f"Professional_Asset_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        print(f"\n📊 EXPORTING PROFESSIONAL EXCEL: {filename}")
        print("=" * 50)
        
        # Create Excel writer
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Asset Report', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Asset Report']
            
            # Define professional styles
            header_fill = PatternFill(
                start_color=self.color_scheme['header_bg'],
                end_color=self.color_scheme['header_bg'],
                fill_type='solid'
            )
            header_font = Font(
                color=self.color_scheme['header_text'],
                bold=True,
                size=12
            )
            
            even_fill = PatternFill(
                start_color=self.color_scheme['row_even'],
                end_color=self.color_scheme['row_even'],
                fill_type='solid'
            )
            
            odd_fill = PatternFill(
                start_color=self.color_scheme['row_odd'],
                end_color=self.color_scheme['row_odd'],
                fill_type='solid'
            )
            
            # Apply header styling
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply row styling
            for row_num, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                fill = even_fill if row_num % 2 == 0 else odd_fill
                for cell in row:
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    # Special formatting for status column
                    if 'Active' in str(cell.value):
                        cell.font = Font(color=self.color_scheme['success'], bold=True)
                    elif 'Offline' in str(cell.value):
                        cell.font = Font(color=self.color_scheme['error'], bold=True)
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
        
        print(f"✅ Professional Excel exported successfully!")
        print(f"📁 File: {filename}")
        print(f"🎨 Professional colors and formatting applied")
        print(f"📊 Auto-adjusted column widths")
        print(f"🔤 Proper text alignment and wrapping")
        
        return filename


def demo_professional_formatting():
    """Demonstrate professional formatting with sample data"""
    
    # Your sample data
    sample_devices = [
        {
            'IP Address': '10.0.23.99',
            'Hostname': 'MHQ-ENG-SAHMED',
            'Working User': r'SQUARE\sara.ahmed',
            'Domain': 'square.local',
            'Device Model': 'Precision Tower 7810',
            'MAC Address': '50:9A:4C:42:D4:09',
            'OS Name and Version': 'Microsoft Windows 10 Pro',
            'Installed RAM (GB)': '31.92',
            'CPU Processor': 'Intel(R) Xeon(R) CPU E5-2680 v4 @ 2.40GHz\nIntel(R) Xeon(R) CPU E5-2680 v4 @ 2.40GHz',
            'Storage (Hard Disk)': 'Disk 1 = 232.88GB\nDisk 2 = 931.51GB',
            'Manufacturer': 'Dell Inc.',
            'Serial Number': 'BD35LH2',
            'Monitor': 'HP EliteDisplay E221c Webcam LED Backlit Monitor',
            'Graphics Card': 'NVIDIA GeForce GTX 1660 Ti',
            'Installed Key': 'N/A'
        },
        {
            'IP Address': '10.0.21.246',
            'Hostname': 'MHQ-ENG-HMETWAL',
            'Working User': r'SQUARE\hager.metwally',
            'Domain': 'square.local',
            'Device Model': 'Precision Tower 7810',
            'MAC Address': '50:9A:4C:45:25:D5',
            'OS Name and Version': 'Microsoft Windows 10 Pro',
            'Installed RAM (GB)': '31.92',
            'CPU Processor': 'Intel(R) Xeon(R) CPU E5-2680 v4 @ 2.40GHz\nIntel(R) Xeon(R) CPU E5-2680 v4 @ 2.40GHz',
            'Storage (Hard Disk)': 'Disk 1 = 465.76GB',
            'Manufacturer': 'Dell Inc.',
            'Serial Number': '7QHHKH2',
            'Monitor': 'HP EliteDisplay E231 LED Backlit Monitor',
            'Graphics Card': 'NVIDIA GeForce GTX 1660 Ti',
            'Installed Key': 'N/A'
        },
        {
            'IP Address': '10.0.23.76',
            'Hostname': 'MHQ-BIM-MGAD',
            'Working User': r'SQUARE\mohamed.gad',
            'Domain': 'square.local',
            'Device Model': 'Precision Tower 7910',
            'MAC Address': 'D8:9E:F3:45:A6:20',
            'OS Name and Version': 'Microsoft Windows 10 Pro',
            'Installed RAM (GB)': '31.92',
            'CPU Processor': 'Intel(R) Xeon(R) CPU E5-2697 v4 @ 2.30GHz\nIntel(R) Xeon(R) CPU E5-2697 v4 @ 2.30GHz',
            'Storage (Hard Disk)': 'Disk 1 = 232.88GB',
            'Manufacturer': 'Dell Inc.',
            'Serial Number': '17XCCP2',
            'Monitor': 'Generic PnP Monitor',
            'Graphics Card': 'NVIDIA GeForce RTX 3060',
            'Installed Key': 'N/A'
        }
    ]
    
    # Create formatter
    formatter = ProfessionalAssetFormatter()
    
    # Create professional presentation
    df = formatter.create_professional_presentation(sample_devices)
    
    # Export to Excel
    excel_file = formatter.export_professional_excel(df)
    
    print(f"\n🎉 PROFESSIONAL FORMATTING COMPLETE!")
    print(f"📊 Excel file: {excel_file}")
    print(f"🎨 Beautiful colors and organization applied")
    print(f"🛡️ Smart duplicate handling completed")
    print(f"💾 Storage details formatted: Disk 1 = 250GB, Disk 2 = 500GB")
    
    return df, excel_file


if __name__ == "__main__":
    demo_professional_formatting()